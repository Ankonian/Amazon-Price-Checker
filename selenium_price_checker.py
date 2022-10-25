from multiprocessing.connection import wait
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import selenium
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl import Workbook
import time
from tkinter import *


#variables to track collected data 
keywords = []
productName = []
productPrice = []
productRating = []
numofProductRating = []
productLink = []
wb = Workbook()


#starting GUI window
root = Tk()
root.geometry("750x250")

#function to save all keywords entered
def saveKeywords():
    newKeyword = entryBox.get()
    if newKeyword != '':
        keywords.append(newKeyword)
    entryBox.delete(0, END)

#saving the search result to product_info spreadsheet
def save_search_to_spreadsheet(keyword, names, prices, ratings, num_of_ratings, links):
    current_sheet = wb.create_sheet(keyword)
    current_sheet.append(["Product Name", "Price", "Star rating", "Number of Ratings", "Product Link"])
    for x in range(0, len(names)):
        current_sheet.cell(row = 2 + x, column = 1).value = names[x]
    for x in range(0, len(prices)):
        current_sheet.cell(row = 2 + x, column = 2).value = prices[x]
    for x in range(0, len(ratings)):
        current_sheet.cell(row = 2 + x, column = 3).value = ratings[x]
    for x in range(0, len(num_of_ratings)):
        current_sheet.cell(row = 2 + x, column = 4).value = num_of_ratings[x]
    for x in range(0, len(links)):
        current_sheet.cell(row = 2 + x, column = 5).value = links[x]
    wb.save('product_infos.xlsx')

#function to search the stored keyword on Amazon
def searchKeywords():
    keywordsLabel = Label(root, text=keywords)
    keywordsLabel.pack()
    driver = webdriver.Chrome(ChromeDriverManager().install())
    
    for keyword in keywords:
        print()
        print("Currently searching for " + keyword)
        print()
        driver.get("https://www.amazon.com/")
        amazonTextBox = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "twotabsearchtextbox"))
            )
        amazonTextBox.send_keys(keyword)
        amazonTextBox.send_keys(Keys.ENTER)
        items = WebDriverWait(driver,10).until(
            EC.presence_of_all_elements_located((By.XPATH, '//div[contains(@class, "s-result-item s-asin")]'))
            )
        for item in items:

            #grabbing item name
            #The try statement is here because Amazon decides to be fancy on random products and user different css styles, thus we need to try different locators
            try:
                itemName = item.find_element(By.XPATH, './/span[@class="a-size-medium a-color-base a-text-normal"]')
                productName.append(itemName.text)
            except:
                itemName = item.find_element(By.XPATH, './/span[@class="a-size-base-plus a-color-base a-text-normal"]')
                productName.append(itemName.text)

            #grabbing item price
            fullPrice = 0
            try:
                wholePrice = item.find_element(By.XPATH, './/span[@class="a-price-whole"]')
                fractionPrice = item.find_element(By.XPATH, './/span[@class="a-price-fraction"]')
            except:
                fullPrice = "not shown on search"
            if(wholePrice != [] and fractionPrice != [] and fullPrice != "not shown on search"):
                fullPrice = wholePrice.text + "." + fractionPrice.text
            else:
                fullPrice = "not shown on search"
            productPrice.append(fullPrice)

            #grabbing item rating and number of ratings
            starRating = 0
            numofRating = 0
            ratingInfos  = item.find_elements(By.XPATH, './/div[@class="a-row a-size-small"]/span')
            if(ratingInfos != []):
                starRating = ratingInfos[0].get_attribute('aria-label')
                #numOfRating = ratingInfos[1].get_attribute('aria-label')
            else:
                starRating = 0
                numofRating = 0
            productRating.append(starRating)
            for x in ratingInfos:
                if x.text != '':
                    numofProductRating.append(x.text)

            #grabbing product link
            link = item.find_element(By.XPATH, './/a[@class="a-link-normal s-underline-text s-underline-link-text s-link-style a-text-normal"]').get_attribute("href")
            productLink.append(link)

        #call the save function to save search results to spreadsheet
        save_search_to_spreadsheet(keyword, productName, productPrice, productRating, numofProductRating, productLink)

    #clearing the saved keywords after searching and saving is complete
    keywords = []

    #remove the empty sheet
    if('Sheet' in wb.sheetnames):
        wb.remove(wb['Sheet'])
        wb.save('product_infos.xlsx')
    

introLabel = Label(root, text="Enter a keyword you wish to search on Amazon:")
introLabel.grid(row=0, column=0)
introLabel.pack()

entryBox = Entry(root, width = 50)
entryBox.pack()

saveKeywordButton = Button(root, text = "Save keyword", command=saveKeywords)
saveKeywordButton.pack()

searchKeywordsButton = Button(root, text = "Search", command=searchKeywords)
searchKeywordsButton.pack()

root.mainloop()

