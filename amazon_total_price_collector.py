import openpyxl

path = "C:/Users/LiLang/Documents/GitHub Repos/Amazon Price Checker/Amazon-Purchase-Automater/item_prices.xlsx"

item_price = openpyxl.load_workbook(path)
sheet1 = item_price.active

for x in range(2, 7):
    cell_obj = sheet1.cell(row=x, column=1)
    print(cell_obj.value)

price = '$274.1'
write_cell = sheet1.cell(row=2, column=2)
write_cell.value = price

item_price.save("C:/Users/LiLang/Documents/GitHub Repos/Amazon Price Checker/Amazon-Purchase-Automater/item_prices.xlsx")